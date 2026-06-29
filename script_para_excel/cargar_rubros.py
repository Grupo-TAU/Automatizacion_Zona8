#!/usr/bin/env python3
"""
Generador de Rubros — Zona 8

Lee el Parte Diario mensual (hojas con nombre AAAAMMDD) y genera un Excel
con la hoja Rubros: N° Problema | Direccion | D1..D7 | E1..E7 | F1..F5 | No Corresponde.

Puede usarse con GUI (doble clic) o desde la terminal:
    python cargar_rubros.py "ruta/al/PD.xlsx"
"""

import os
import re
import sys
import threading
import tkinter as tk
from collections import Counter, defaultdict
from tkinter import filedialog, messagebox, scrolledtext

from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------

RUTA_DEFAULT = ""

CODIGOS = [
    "D1", "D2", "D3", "D4", "D5", "D6", "D7",
    "E1", "E2", "E3", "E4", "E5", "E6", "E7",
    "F1", "F2", "F3", "F4", "F5",
    "No Corresponde",
]

COL_PROBLEMA = 1   # A
COL_DIRECCION = 3  # C
COL_RUBRO = 8      # H

HOJA_FECHA = re.compile(r"^\d{8}$")
RE_CODIGO = re.compile(r"^\s*(D[1-7]|E[1-7]|F[1-5]|No\s+Corresponde)", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Logica principal
# ---------------------------------------------------------------------------

def _normalizar_codigo(raw):
    if len(raw) == 2:
        return raw.upper()
    return "No Corresponde"


def recolectar_rubros(path_parte, log=print):
    wb = load_workbook(path_parte, data_only=True)
    rubros = defaultdict(lambda: {'direccion': '', 'conteos': Counter()})
    sin_numero = defaultdict(lambda: {'conteos': Counter()})
    descartados = []

    hojas = [n for n in wb.sheetnames if HOJA_FECHA.match(n)]
    log(f"Hojas de fecha encontradas: {len(hojas)}")

    for nombre in hojas:
        log(f"  Procesando hoja {nombre}...")
        ws = wb[nombre]
        for r in range(1, ws.max_row + 1):
            rubro = ws.cell(r, COL_RUBRO).value
            if not rubro:
                continue
            m = RE_CODIGO.match(str(rubro))
            if not m:
                prob = ws.cell(r, COL_PROBLEMA).value
                if isinstance(prob, (int, float)):
                    descartados.append((nombre, str(int(prob)), str(rubro)[:40]))
                continue
            prob = ws.cell(r, COL_PROBLEMA).value
            direccion = ws.cell(r, COL_DIRECCION).value
            codigo = _normalizar_codigo(m.group(1))
            if not isinstance(prob, (int, float)):
                if direccion:
                    sin_numero[str(direccion)]['conteos'][codigo] += 1
                continue
            n = str(int(prob))
            if not rubros[n]['direccion'] and direccion:
                rubros[n]['direccion'] = str(direccion)
            rubros[n]['conteos'][codigo] += 1

    wb.close()
    return (
        {n: v for n, v in rubros.items() if v['conteos']},
        dict(sin_numero),
        descartados,
    )


def generar_excel(path_parte, log=print):
    """Procesa el Parte Diario y escribe el archivo Rubro_D. Devuelve la ruta del archivo generado."""
    rubros, sin_numero, descartados = recolectar_rubros(path_parte, log=log)
    log(f"\nProblemas con rubro valido: {len(rubros)}")
    log(f"Filas sin N° Problema (solo ubicacion): {len(sin_numero)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Rubros"
    ws.append(["N° Problema", "Direccion"] + CODIGOS)

    for n, datos in rubros.items():
        fila = [n, datos['direccion']]
        for codigo in CODIGOS:
            fila.append(datos['conteos'].get(codigo) or None)
        ws.append(fila)

    for direccion, datos in sin_numero.items():
        fila = [None, direccion]
        for codigo in CODIGOS:
            fila.append(datos['conteos'].get(codigo) or None)
        ws.append(fila)

    salida = path_parte.replace(".xlsx", "_Rubros.xlsx")
    wb.save(salida)

    log(f"\nFilas generadas: {len(rubros) + len(sin_numero)}")
    if descartados:
        log(f"Ignorados (rubro sin codigo valido): {len(descartados)}")
        for hoja, n, txt in descartados:
            log(f"  [{hoja}] Problema {n}: {txt}")
    log(f"\nGuardado en:\n{salida}")
    return salida


# ---------------------------------------------------------------------------
# Interfaz grafica
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Generador de Rubros — Zona 8")
        self.resizable(False, False)
        self._salida = None
        self._build()

    def _build(self):
        self.configure(padx=16, pady=12)

        # ── Selector de archivo ──────────────────────────────────────────
        tk.Label(self, text="Archivo Parte Diario:", anchor="w").grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 2)
        )
        self.var_path = tk.StringVar(value=RUTA_DEFAULT)
        tk.Entry(self, textvariable=self.var_path, width=62).grid(
            row=1, column=0, sticky="ew"
        )
        tk.Button(self, text="Examinar…", command=self._examinar).grid(
            row=1, column=1, padx=(6, 0)
        )

        # ── Boton principal ──────────────────────────────────────────────
        self.btn_generar = tk.Button(
            self,
            text="Generar Rubros",
            command=self._generar,
            bg="#0078d4",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            padx=24,
            pady=8,
            relief="flat",
            cursor="hand2",
        )
        self.btn_generar.grid(row=2, column=0, columnspan=2, pady=14)

        # ── Log ──────────────────────────────────────────────────────────
        tk.Label(self, text="Registro:", anchor="w").grid(
            row=3, column=0, columnspan=2, sticky="w", pady=(0, 2)
        )
        self.log_area = scrolledtext.ScrolledText(
            self, width=72, height=16, state="disabled",
            font=("Consolas", 9), bg="#f5f5f5"
        )
        self.log_area.grid(row=4, column=0, columnspan=2)

        # ── Boton abrir resultado ────────────────────────────────────────
        self.btn_abrir = tk.Button(
            self,
            text="Abrir archivo generado",
            command=self._abrir,
            state="disabled",
            cursor="hand2",
        )
        self.btn_abrir.grid(row=5, column=0, columnspan=2, pady=(10, 0))

    # ── Acciones ─────────────────────────────────────────────────────────

    def _examinar(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Parte Diario",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos los archivos", "*.*")],
        )
        if path:
            self.var_path.set(path)

    def _generar(self):
        path = self.var_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning(
                "Archivo no encontrado",
                "El archivo indicado no existe.\nVerificá la ruta o usá 'Examinar' para buscarlo.",
            )
            return

        self.btn_generar.config(state="disabled", text="Procesando…")
        self.btn_abrir.config(state="disabled")
        self._limpiar_log()
        self._salida = None

        def tarea():
            try:
                salida = generar_excel(path, log=self._log_desde_hilo)
                self._salida = salida
                self.after(0, self._on_exito)
            except Exception as exc:
                self.after(0, lambda e=exc: self._on_error(str(e)))

        threading.Thread(target=tarea, daemon=True).start()

    def _abrir(self):
        if self._salida and os.path.exists(self._salida):
            os.startfile(self._salida)

    # ── Helpers de log ───────────────────────────────────────────────────

    def _log(self, msg):
        self.log_area.config(state="normal")
        self.log_area.insert("end", msg + "\n")
        self.log_area.see("end")
        self.log_area.config(state="disabled")

    def _log_desde_hilo(self, msg):
        self.after(0, lambda m=msg: self._log(m))

    def _limpiar_log(self):
        self.log_area.config(state="normal")
        self.log_area.delete("1.0", "end")
        self.log_area.config(state="disabled")

    # ── Callbacks de resultado ───────────────────────────────────────────

    def _on_exito(self):
        self.btn_generar.config(state="normal", text="Generar Rubros")
        self.btn_abrir.config(state="normal")
        messagebox.showinfo("Listo", f"Archivo generado correctamente:\n\n{self._salida}")

    def _on_error(self, msg):
        self.btn_generar.config(state="normal", text="Generar Rubros")
        self._log(f"\n[ERROR] {msg}")
        messagebox.showerror("Error al procesar", f"Ocurrió un error:\n\n{msg}")


# ---------------------------------------------------------------------------
# Entrada
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) > 1:
        # Modo terminal: python cargar_rubros.py ruta.xlsx
        path = sys.argv[1].strip()
        if not os.path.exists(path):
            print(f"Archivo no encontrado: {path}")
            sys.exit(1)
        generar_excel(path)
    else:
        # Modo GUI
        App().mainloop()


if __name__ == "__main__":
    main()
